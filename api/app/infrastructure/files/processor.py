"""
Converts uploaded raw base64 files into stored attachment dicts.
Images are kept as base64 for vision. Documents have text extracted.
"""
from __future__ import annotations

import base64
import io

_MIME_IMAGE = {"image/jpeg", "image/png", "image/gif", "image/webp"}


def process_attachment(filename: str, mime_type: str, data_b64: str) -> dict:
    """Return a stored attachment dict for one uploaded file."""
    if mime_type in _MIME_IMAGE:
        return {"type": "image", "filename": filename, "mime_type": mime_type, "data": data_b64}

    raw = base64.b64decode(data_b64)
    if mime_type == "application/pdf":
        text = _extract_pdf(raw)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        text = _extract_docx(raw)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        text = _extract_xlsx(raw)
    else:
        try:
            text = raw.decode("utf-8", errors="replace")
        except Exception:
            text = raw.decode("latin-1", errors="replace")

    return {"type": "document", "filename": filename, "mime_type": mime_type, "extracted_text": text}


def _extract_pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        return "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    except Exception as e:
        return f"[PDF 提取失败: {e}]"


def _extract_docx(raw: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs if p.text).strip()
    except Exception as e:
        return f"[DOCX 提取失败: {e}]"


def _extract_xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        parts: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f"[Sheet: {name}]")
            for row in ws.iter_rows(values_only=True):
                parts.append("\t".join("" if c is None else str(c) for c in row))
        return "\n".join(parts).strip()
    except Exception as e:
        return f"[XLSX 提取失败: {e}]"
